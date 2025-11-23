"""
Helper utilities for URDB Tariff Viewer.

This module contains common utility functions used throughout the application.
"""

from typing import Any, Optional, Union
import re
from datetime import datetime, timedelta
import pandas as pd
import io


def format_currency(amount: Union[int, float], precision: int = 2) -> str:
    """
    Format a number as currency.
    
    Args:
        amount (Union[int, float]): Amount to format
        precision (int): Number of decimal places
        
    Returns:
        str: Formatted currency string
    """
    if amount is None:
        return "$0.00"
    
    return f"${amount:,.{precision}f}"


def format_percentage(value: Union[int, float], precision: int = 1) -> str:
    """
    Format a decimal as a percentage.
    
    Args:
        value (Union[int, float]): Decimal value (0.5 = 50%)
        precision (int): Number of decimal places
        
    Returns:
        str: Formatted percentage string
    """
    if value is None:
        return "0.0%"
    
    return f"{value * 100:.{precision}f}%"


def get_month_name(month_index: int, abbreviated: bool = True) -> str:
    """
    Get month name from index.
    
    Args:
        month_index (int): Month index (0-11 or 1-12)
        abbreviated (bool): Whether to return abbreviated name
        
    Returns:
        str: Month name
    """
    full_months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    
    abbrev_months = [
        'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
    ]
    
    # Handle both 0-based and 1-based indexing
    if month_index > 12:
        month_index = month_index % 12
    elif month_index > 11:
        month_index -= 1
    
    # Clamp to valid range
    month_index = max(0, min(11, month_index))
    
    if abbreviated:
        return abbrev_months[month_index]
    else:
        return full_months[month_index]


def clean_filename(filename: str) -> str:
    """
    Clean a filename by removing invalid characters.
    
    Args:
        filename (str): Original filename
        
    Returns:
        str: Cleaned filename
    """
    # Remove invalid characters
    cleaned = re.sub(r'[<>:"/\\|?*]', '_', filename)
    
    # Remove multiple underscores
    cleaned = re.sub(r'_+', '_', cleaned)
    
    # Remove leading/trailing underscores and dots
    cleaned = cleaned.strip('_.')
    
    # Ensure it's not empty
    if not cleaned:
        cleaned = "untitled"
    
    return cleaned


def safe_float_conversion(value: Any, default: float = 0.0) -> float:
    """
    Safely convert a value to float.
    
    Args:
        value (Any): Value to convert
        default (float): Default value if conversion fails
        
    Returns:
        float: Converted value or default
    """
    if value is None:
        return default
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int_conversion(value: Any, default: int = 0) -> int:
    """
    Safely convert a value to integer.
    
    Args:
        value (Any): Value to convert
        default (int): Default value if conversion fails
        
    Returns:
        int: Converted value or default
    """
    if value is None:
        return default
    
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def truncate_text(text: str, max_length: int = 50, suffix: str = "...") -> str:
    """
    Truncate text to a maximum length.
    
    Args:
        text (str): Text to truncate
        max_length (int): Maximum length
        suffix (str): Suffix to add when truncated
        
    Returns:
        str: Truncated text
    """
    if len(text) <= max_length:
        return text
    
    return text[:max_length - len(suffix)] + suffix


def validate_email(email: str) -> bool:
    """
    Validate an email address format.
    
    Args:
        email (str): Email address to validate
        
    Returns:
        bool: True if valid email format
    """
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def format_file_size(size_bytes: int) -> str:
    """
    Format file size in human readable format.
    
    Args:
        size_bytes (int): Size in bytes
        
    Returns:
        str: Formatted size string
    """
    if size_bytes == 0:
        return "0 B"
    
    units = ["B", "KB", "MB", "GB", "TB"]
    unit_index = 0
    size = float(size_bytes)
    
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    
    if unit_index == 0:
        return f"{int(size)} {units[unit_index]}"
    else:
        return f"{size:.1f} {units[unit_index]}"


def get_current_timestamp() -> str:
    """
    Get current timestamp as string.
    
    Returns:
        str: Current timestamp in YYYY-MM-DD HH:MM:SS format
    """
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_timestamp(timestamp_str: str) -> Optional[datetime]:
    """
    Parse a timestamp string into datetime object.
    
    Args:
        timestamp_str (str): Timestamp string
        
    Returns:
        Optional[datetime]: Parsed datetime or None if parsing fails
    """
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y"
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(timestamp_str, fmt)
        except ValueError:
            continue
    
    return None


def calculate_percentage_change(old_value: float, new_value: float) -> float:
    """
    Calculate percentage change between two values.
    
    Args:
        old_value (float): Original value
        new_value (float): New value
        
    Returns:
        float: Percentage change (as decimal, e.g., 0.1 = 10% increase)
    """
    if old_value == 0:
        return 1.0 if new_value > 0 else 0.0
    
    return (new_value - old_value) / old_value


def deep_merge_dicts(dict1: dict, dict2: dict) -> dict:
    """
    Deep merge two dictionaries.
    
    Args:
        dict1 (dict): First dictionary
        dict2 (dict): Second dictionary (takes precedence)
        
    Returns:
        dict: Merged dictionary
    """
    result = dict1.copy()
    
    for key, value in dict2.items():
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            result[key] = deep_merge_dicts(result[key], value)
        else:
            result[key] = value
    
    return result


def generate_energy_rates_excel(tariff_viewer, year: int = 2025) -> bytes:
    """
    Generate an Excel file with multiple sheets containing energy and demand rate data.
    
    Args:
        tariff_viewer: TariffViewer instance with energy and demand rate data
        year (int): Year for timeseries generation (default 2025)
        
    Returns:
        bytes: Excel file content as bytes
    """
    from openpyxl.styles import numbers
    
    # Create a bytes buffer to hold the Excel file
    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # ENERGY RATES SHEETS
        
        # Sheet 1: Current Energy Rate Table (TOU Labels)
        try:
            tou_table = tariff_viewer.create_tou_labels_table()
            if not tou_table.empty:
                # Convert formatted strings to numeric for Excel
                excel_tou = tou_table.copy()
                for col in ['Base Rate ($/kWh)', 'Adjustment ($/kWh)', 'Total Rate ($/kWh)']:
                    if col in excel_tou.columns:
                        excel_tou[col] = excel_tou[col].str.replace('$', '').astype(float)
                if '% of Year' in excel_tou.columns:
                    excel_tou['% of Year'] = excel_tou['% of Year'].str.replace('%', '').astype(float) / 100
                
                excel_tou.to_excel(writer, sheet_name='Energy Rate Table', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['Energy Rate Table']
                headers = list(excel_tou.columns)
                for col_idx, col_name in enumerate(headers, start=1):
                    if col_name in ['Base Rate ($/kWh)', 'Adjustment ($/kWh)', 'Total Rate ($/kWh)']:
                        for row_idx in range(2, len(excel_tou) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
                    elif col_name == '% of Year':
                        for row_idx in range(2, len(excel_tou) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.number_format = '0.0%'
            else:
                # Create a note if no energy rates
                no_data_df = pd.DataFrame({'Note': ['No energy rate structure found in tariff']})
                no_data_df.to_excel(writer, sheet_name='Energy Rate Table', index=False)
        except Exception as e:
            # Create an empty sheet with error message
            error_df = pd.DataFrame({'Error': [f'Could not generate rate table: {str(e)}']})
            error_df.to_excel(writer, sheet_name='Energy Rate Table', index=False)
        
        # Sheet 2: Weekday Energy Rates Heatmap
        weekday_df = tariff_viewer.weekday_df.copy()
        weekday_df.columns = [f'{h:02d}:00' for h in range(24)]
        weekday_df.to_excel(writer, sheet_name='Weekday Energy Rates')
        
        # Apply accounting format to rate values
        worksheet = writer.sheets['Weekday Energy Rates']
        for row_idx in range(2, len(weekday_df) + 2):
            for col_idx in range(2, len(weekday_df.columns) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
        
        # Sheet 3: Weekend Energy Rates Heatmap
        weekend_df = tariff_viewer.weekend_df.copy()
        weekend_df.columns = [f'{h:02d}:00' for h in range(24)]
        weekend_df.to_excel(writer, sheet_name='Weekend Energy Rates')
        
        # Apply accounting format to rate values
        worksheet = writer.sheets['Weekend Energy Rates']
        for row_idx in range(2, len(weekend_df) + 2):
            for col_idx in range(2, len(weekend_df.columns) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
        
        # Sheet 4: Full Year Energy Timeseries
        timeseries_df = generate_energy_rate_timeseries(tariff_viewer, year)
        timeseries_df.to_excel(writer, sheet_name='Energy Timeseries', index=False)
        
        # Apply format to energy rate column
        if 'energy_rate_$/kWh' in timeseries_df.columns:
            worksheet = writer.sheets['Energy Timeseries']
            rate_col_idx = list(timeseries_df.columns).index('energy_rate_$/kWh') + 1
            for row_idx in range(2, len(timeseries_df) + 2):
                cell = worksheet.cell(row=row_idx, column=rate_col_idx)
                cell.number_format = '_($* #,##0.0000_);_($* (#,##0.0000);_($* "-"????_);_(@_)'
        
        # DEMAND RATES SHEETS
        
        # Sheet 5: Current Demand Rate Table (Demand Labels)
        try:
            demand_table = tariff_viewer.create_demand_labels_table()
            if not demand_table.empty:
                # Convert formatted strings to numeric for Excel
                excel_demand = demand_table.copy()
                for col in ['Base Rate ($/kW)', 'Adjustment ($/kW)', 'Total Rate ($/kW)']:
                    if col in excel_demand.columns:
                        excel_demand[col] = excel_demand[col].str.replace('$', '').astype(float)
                if '% of Year' in excel_demand.columns:
                    excel_demand['% of Year'] = excel_demand['% of Year'].str.replace('%', '').astype(float) / 100
                
                excel_demand.to_excel(writer, sheet_name='Demand Rate Table', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['Demand Rate Table']
                headers = list(excel_demand.columns)
                for col_idx, col_name in enumerate(headers, start=1):
                    if col_name in ['Base Rate ($/kW)', 'Adjustment ($/kW)', 'Total Rate ($/kW)']:
                        for row_idx in range(2, len(excel_demand) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                    elif col_name == '% of Year':
                        for row_idx in range(2, len(excel_demand) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.number_format = '0.0%'
            else:
                # Create a note if no demand rates
                no_data_df = pd.DataFrame({'Note': ['No demand rate structure found in tariff']})
                no_data_df.to_excel(writer, sheet_name='Demand Rate Table', index=False)
        except Exception as e:
            # Create an empty sheet with error message
            error_df = pd.DataFrame({'Error': [f'Could not generate demand rate table: {str(e)}']})
            error_df.to_excel(writer, sheet_name='Demand Rate Table', index=False)
        
        # Sheet 6: Weekday Demand Rates Heatmap
        demand_weekday_df = tariff_viewer.demand_weekday_df.copy()
        demand_weekday_df.columns = [f'{h:02d}:00' for h in range(24)]
        demand_weekday_df.to_excel(writer, sheet_name='Weekday Demand Rates')
        
        # Apply accounting format to rate values
        worksheet = writer.sheets['Weekday Demand Rates']
        for row_idx in range(2, len(demand_weekday_df) + 2):
            for col_idx in range(2, len(demand_weekday_df.columns) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Sheet 7: Weekend Demand Rates Heatmap
        demand_weekend_df = tariff_viewer.demand_weekend_df.copy()
        demand_weekend_df.columns = [f'{h:02d}:00' for h in range(24)]
        demand_weekend_df.to_excel(writer, sheet_name='Weekend Demand Rates')
        
        # Apply accounting format to rate values
        worksheet = writer.sheets['Weekend Demand Rates']
        for row_idx in range(2, len(demand_weekend_df) + 2):
            for col_idx in range(2, len(demand_weekend_df.columns) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        
        # Sheet 8: Flat Demand Rates (Seasonal/Monthly)
        flat_demand_df = tariff_viewer.flat_demand_df.copy()
        flat_demand_df.to_excel(writer, sheet_name='Flat Demand Rates')
        
        # Apply accounting format to rate values
        worksheet = writer.sheets['Flat Demand Rates']
        for row_idx in range(2, len(flat_demand_df) + 2):
            for col_idx in range(2, len(flat_demand_df.columns) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Get the Excel file content as bytes
    excel_data = output.getvalue()
    return excel_data


def generate_energy_rate_timeseries(tariff_viewer, year: int = 2025) -> pd.DataFrame:
    """
    Generate a full year of timeseries data with timestamps and corresponding energy rates.
    
    Args:
        tariff_viewer: TariffViewer instance with energy rate data
        year (int): Year for timeseries generation
        
    Returns:
        pd.DataFrame: DataFrame with 'timestamp' and 'energy_rate_$/kWh' columns
    """
    # Generate timestamps for full year at 15-minute intervals
    start_date = datetime(year, 1, 1, 0, 0, 0)
    end_date = datetime(year, 12, 31, 23, 45, 0)
    
    # Create timestamp range (15-minute intervals)
    timestamps = []
    current_time = start_date
    while current_time <= end_date:
        timestamps.append(current_time)
        current_time += timedelta(minutes=15)
    
    # Create DataFrame
    df = pd.DataFrame({'timestamp': timestamps})
    
    # Add derived columns for rate lookup
    df['month'] = df['timestamp'].dt.month - 1  # 0-indexed for array lookup
    df['hour'] = df['timestamp'].dt.hour
    df['weekday'] = df['timestamp'].dt.weekday  # 0=Monday, 6=Sunday
    df['is_weekend'] = df['weekday'] >= 5  # Saturday=5, Sunday=6
    
    # Get the appropriate rate for each timestamp
    energy_rates = []
    
    for idx, row in df.iterrows():
        month_idx = row['month']
        hour = row['hour']
        is_weekend = row['is_weekend']
        
        # Select appropriate rate DataFrame
        if is_weekend:
            rate_df = tariff_viewer.weekend_df
        else:
            rate_df = tariff_viewer.weekday_df
        
        # Get rate from DataFrame (month is row index, hour is column)
        try:
            rate = rate_df.iloc[month_idx, hour]
            energy_rates.append(rate)
        except (IndexError, KeyError):
            energy_rates.append(0.0)
    
    # Create final DataFrame with only timestamp and rate
    result_df = pd.DataFrame({
        'timestamp': df['timestamp'],
        'energy_rate_$/kWh': energy_rates
    })
    
    return result_df
